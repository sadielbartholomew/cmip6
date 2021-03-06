"""
.. module:: generate_cim.py
   :license: GPL/CeCIL
   :platform: Unix, Windows
   :synopsis: Generates CMIP6 JSON documents from XLS files.

.. moduleauthor:: Mark Conway-Greenslade <momipsl@ipsl.jussieu.fr>

"""
import argparse
import collections
import json
import os

import openpyxl

from lib.utils import io_mgr
from lib.utils import logger
from lib.utils import vocabs


# Define command line argument parser.
_ARGS = argparse.ArgumentParser("Generates CMIP6 responsible parties JSON files.")
_ARGS.add_argument(
    "--institution-id",
    help="An institution identifier",
    dest="institution_id",
    type=str
    )



def _main(args):
    """Main entry point.

    """
    for i in vocabs.get_institutes(args.institution_id):
        _write(i)


def _write(i):
    """Writes parties JSON file for a particular institute.

    """
    try:
        spreadsheet = _get_spreadsheet(i)
    except IOError:
        msg = '{} parties spreadsheet not found'.format(i.canonical_name)
        logger.log_warning(msg)
    else:
        content = _get_content(i, spreadsheet)
        if content:
            _write_content(i, content)


def _get_spreadsheet(i):
    """Returns a spreadsheet for processing.

    """
    path = io_mgr.get_parties_spreadsheet(i)
    if not os.path.exists(path):
        raise IOError()

    return openpyxl.load_workbook(path, read_only=True)


def _get_content(i, spreadsheet):
    """Returns content to be written to file system.

    """
    # Initialise output.
    obj = collections.OrderedDict()
    obj['mipEra'] = "cmip6"
    obj['institute'] = i.canonical_name
    obj['seedingSource'] = 'Spreadsheet'
    obj['content'] = []

    # Process spreadsheet.
    for _, worksheet in [(i, j) for i, j in enumerate(spreadsheet) if i > 1]:
        _set_xls_content(obj, worksheet)

    return obj


def _write_content(i, content):
    """Writes json content to file system.

    """
    fpath = io_mgr.get_parties_json(i)
    with open(fpath, 'w') as fstream:
        fstream.write(json.dumps(content, indent=4))


def _set_xls_content(obj, worksheet):
    """Sets content for a particular specialization.

    """
    for row in worksheet.iter_rows(min_row=3, max_col=7, max_row=worksheet.max_row):
        if not row:
            continue
        if row[0].value is None:
            continue

        mnemonic, name, is_organisation, address_postal, \
        address_email, url, orcid = [i.value for i in row]

        party = collections.OrderedDict()
        party['mnemonic'] = mnemonic
        party['name'] = name
        party['is_organisation'] = is_organisation in ['yes', 'y']
        party['address_postal'] = address_postal
        party['address_email'] = address_email
        party['url'] = url
        party['orcid'] = orcid

        obj['content'].append(party)


# Main entry point.
if __name__ == '__main__':
    _main(_ARGS.parse_args())
