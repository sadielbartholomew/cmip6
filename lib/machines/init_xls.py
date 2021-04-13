"""
.. module:: init_machines_xls.py
   :license: GPL/CeCIL
   :platform: Unix, Windows
   :synopsis: Initialises CMIP6 machines spreadsheets.

.. moduleauthor::
   Mark Conway-Greenslade <momipsl@ipsl.jussieu.fr>
   Sadie Bartholomew <sadie.bartholomew@ncas.ac.uk>

"""
import argparse
import os
import shutil

from copy import copy

from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

from lib.utils import io_mgr, logger, vocabs, constants


# Define command line argument parser.
_ARGS = argparse.ArgumentParser("Initialises CMIP6 machines spreadsheets.")
_ARGS.add_argument(
    "--institution-id",
    help="An institution identifier",
    dest="institution_id",
    type=str,
    default="all"
    )
_ARGS.add_argument(
    "--xls-template",
    help="Path to XLS template",
    dest="xls_template",
    type=str
    )


# Define formatting styles to apply on the generated spreadsheets.
QUESTION_HEADER_STYLE = NamedStyle(name="QUESTION_HEADER_STYLE")
QUESTION_HEADER_STYLE.fill = PatternFill(start_color="337AB7")
QUESTION_HEADER_STYLE.font = Font(
    name="Helvetica Neue", size=14, bold=True, color="FFFFFF")
QUESTION_INPUT_BOX_STYLE = NamedStyle(name="QUESTION_INPUT_BOX_STYLE")
QUESTION_INPUT_BOX_STYLE.fill = PatternFill(start_color="CCCCCC")
QUESTION_INPUT_BOX_STYLE.font = Font(
    name="Helvetica Neue", size=14, bold=False, color="000000")


def copy_cell(sheet, cell_to_copy_to, cell_to_copy_from):
    """Apply the value and background style of a cell to another named cell."""
    sheet[cell_to_copy_to] = sheet[cell_to_copy_from].value
    sheet[cell_to_copy_to]._style = copy(sheet[cell_to_copy_from]._style)


def get_applicable_mips_with_experiments(institution):
    """TODO, and MOVE TO UTILS."""
    # First find all of the applicable MIPs for the institute
    all_applicable_mips = list()
    for source in vocabs.get_institute_sources(institution):
        for mip in source.activity_participation:
            all_applicable_mips.append(mip.encode())

    # Some experiments apply to multiple MIPs, so handle the mapping:
    exps_to_mips = dict()
    for exp in vocabs.get_experiments:
        mips = [mip.encode() for mip in exp.data["activity_id"]]
        exps_to_mips.update({exp.canonical_name.encode(): mips})

    # Now change the mapping from current exp -> MIP to the required MIP -> exp
    mips_to_exps = {}
    for exp, mips in exps_to_mips.items():
        for mip in mips:
            if mip in all_applicable_mips:  # filter out non-applicable
                mips_to_exps.setdefault(mip.encode(), []).append(exp)

    return mips_to_exps


def process_enum_options(experiment_list):
    """TODO."""
    if len(experiment_list) == 1:
        return experiment_list
    elif len(experiment_list) >= 1:
        return ["ALL", "NONE"] + experiment_list
    else:
        raise ValueError(
            "No experiments for this MIP and institute. This should not be "
            "the case so something went wrong with the vocabs processing.")


def set_institute_name_in_xls(institution, spreadsheet):
    """Set the institute name on the frontis and machine tab header."""
    short_name = institution.canonical_name.upper().encode()

    # Write institute name on frontis page.
    frontis_sheet = spreadsheet["Frontis"]
    frontis_sheet["B4"] = "{} ({})".format(
        institution.data["name"].encode(), short_name)  # long name w/ acronym

    # Write header including institute name on the machine tab.
    machines_sheet = spreadsheet["Machine 1"]
    machines_sheet["B1"] = "{} Machine for {}".format(
        short_name, constants.CMIP6_MIP_ERA.upper())


def manage_no_applicable_items_found(
        sheet, start_row, item_missing, question_number):
    """TODO."""
    # Delete then re-insert same number of rows to clear 'question' styling
    sheet.delete_rows(start_row - 1, 2)
    sheet.insert_rows(start_row - 1, 2)

    note_cell_address = "A{}".format(start_row)
    note_cell = sheet[note_cell_address]
    note_cell.font = Font(name="Helvetica Neue", size=14, bold=True)
    sheet[note_cell_address] = "N/A: NO SECTIONS"

    msg_cell_address = "B{}".format(start_row)
    msg_cell = sheet[msg_cell_address]
    msg_cell.font = Font(
        name="Helvetica Neue", size=14, bold=True, color="BE0E0E")  # red
    sheet[msg_cell_address] = (
        "!!! No applicable {} detected for your institute. If that is "
        "correct, please skip this question ({}), otherwise contact "
        "ES-DOC support via the email above so we can rectify "
        "this. !!!".format(item_missing, question_number)
    )
    sheet.row_dimensions[start_row].height = 50
    sheet["B{}".format(start_row)].alignment = Alignment(
        wrap_text=True)


def set_applicable_models_in_xls(institution, spreadsheet):
    """Create a question concerning every CMIP6 model ran by the institute."""
    machines_sheet = spreadsheet["Machine 1"]
    active_row = 390  # last row where model question cell block needs updating
    label_base = "1.8.2.{}"
    section_number = 1  # ready to increment label_base end sectioning from one

    # Merge columns on the B-D for the question block.
    start_merging_at_cell = active_row - 20
    for coor in range(start_merging_at_cell, start_merging_at_cell + 500):
         machines_sheet.merge_cells("B{0}:D{0}".format(coor))
         machines_sheet.row_dimensions[start_merging_at_cell].height = 40

    if not vocabs.get_model_configurations(institution):
        manage_no_applicable_items_found(
            machines_sheet, active_row, "models", "1.8")
        return active_row

    for model in vocabs.get_model_configurations(institution):
        if section_number != 1:
            # Create new block of three rows: two question rows plus one gap
            machines_sheet.insert_rows(active_row + 1, 3)
            for row_number in range(active_row - 2, 3):
                machines_sheet.row_dimensions[row_number].height = 40

            active_row += 3  # move active row down to top-most block to update

            # Style cells first, as this avoids style issues with re-assignment
            for col in ("A", "B"):
                machines_sheet[
                    "{}{}".format(col, active_row - 1)
                ].style = QUESTION_HEADER_STYLE
            machines_sheet[
                "A{0}".format(active_row)
            ].style = QUESTION_INPUT_BOX_STYLE

            # Take four cells as the basic block with the question for a model.
            # Set the newly inserted rows to be identical to the basic block.
            # (Note 'for col in ("A", "B"):' use to prevent duplication breaks
            # the formatting for unknown reasons, so will have to leave as-is!)
            copy_cell(
                machines_sheet,
                "A{}".format(active_row),
                "A{}".format(active_row - 3)
            )
            copy_cell(
                machines_sheet,
                "A{}".format(active_row - 1),
                "A{}".format(active_row - 4)
            )
            copy_cell(
                machines_sheet,
                "B{}".format(active_row),
                "B{}".format(active_row - 3)
            )
            copy_cell(
                machines_sheet,
                "B{}".format(active_row - 1),
                "B{}".format(active_row - 4)
            )

            # Finally, add a 'YES/NO' drop-down box (i.e. some data validation)
            yes_or_no_dropdown = DataValidation(
                type="list", formula1='"YES,NO"', allow_blank=True)
            machines_sheet.add_data_validation(yes_or_no_dropdown)
            yes_or_no_dropdown.add("B{}".format(active_row))

        # Change label cell so final (sub-sub-sub-section) number increments
        question_label = label_base.format(section_number)
        machines_sheet["A{}".format(active_row - 1)] = question_label

        # Process in model name (note assume Python 2 hence encode necessary)
        machines_sheet["B{}".format(active_row - 1)] = model.raw_name.encode()

        section_number += 1  # update sub-sub-sub-section for next question

    return active_row  # so we know where to start with next section


def set_applicable_experiments_in_xls(institution, spreadsheet, start_cell):
    """TODO."""
    machines_sheet = spreadsheet["Machine 1"]
     # Last row where question block needs updating. To find this, grab the
     # final active cell of the model questions which sit above and add 20
     # because there are 20 cells of context before the question list begins.
    active_row = start_cell + 19

    label_base = "1.9.2.{}"
    section_number = 1  # ready to increment label_base end sectioning from one

    # Make cell with instructions much taller as it holds a lot of text
    # (can't set via original template as it messes up processed cell heights).
    machines_sheet.row_dimensions[active_row - 3].height = 220

    # First question: add an 'ALL, SOME' drop-down box
    all_or_some_dropdown = DataValidation(
        type="list", formula1='"ALL, SOME"', allow_blank=True)
    machines_sheet.add_data_validation(all_or_some_dropdown)
    all_or_some_dropdown.add("B{}".format(active_row - 10))

    if not get_applicable_mips_with_experiments(institution):
        manage_no_applicable_items_found(
            machines_sheet, active_row, "MIPs", "1.9")
        return active_row

    # Iterate over MIPs to list and tied experiments for each dropdown box
    exps_dropdown = []  # to register separate data alidation items later
    for mip, exps in get_applicable_mips_with_experiments(institution).items():

        if section_number != 1:
            # Create new block of three rows: two question rows plus one gap
            machines_sheet.insert_rows(active_row + 1, 3)
            for row_number in range(active_row - 2, 3):
                machines_sheet.row_dimensions[row_number].height = 40

            active_row += 3  # move active row down to top-most block to update

            # Style cells first, as this avoids style issues with re-assignment
            for col in ("A", "B"):
                machines_sheet[
                    "{}{}".format(col, active_row - 1)
                ].style = QUESTION_HEADER_STYLE
            machines_sheet[
                "A{0}".format(active_row)
            ].style = QUESTION_INPUT_BOX_STYLE

            # Take four cells as the basic block with the question for a model.
            # Set the newly inserted rows to be identical to the basic block.
            # (Note 'for col in ("A", "B"):' use to prevent duplication breaks
            # the formatting for unknown reasons, so will have to leave as-is!)
            copy_cell(
                machines_sheet,
                "A{}".format(active_row),
                "A{}".format(active_row - 3)
            )
            copy_cell(
                machines_sheet,
                "A{}".format(active_row - 1),
                "A{}".format(active_row - 4)
            )
            copy_cell(
                machines_sheet,
                "B{}".format(active_row),
                "B{}".format(active_row - 3)
            )
            copy_cell(
                machines_sheet,
                "B{}".format(active_row - 1),
                "B{}".format(active_row - 4)
            )

        # Add a drop-down box listing all experiments else text of the only one
        options = process_enum_options(exps)
        if len(options) == 1:
            # Set text naming the only experiment, no need for a drop-down box
            machines_sheet["B{}".format(active_row)] = options[0]
        else:
            # First clear,  else dropdown-box is pre-selected w/ random choice.
            machines_sheet["B{}".format(active_row)] = ""

            # Add the drop-down box options for ALL, NONE and the experiments
            options_formula = ",".join(options)
            exps_dropdown.append(
                DataValidation(
                    type="list",
                    formula1='"{}"'.format(options_formula),
                    allow_blank=True
                )
            )
            machines_sheet.add_data_validation(exps_dropdown[-1])
            exps_dropdown[-1].add("B{}".format(active_row))

        # Change label cell so final (sub-sub-sub-section) number increments
        question_label = label_base.format(section_number)
        machines_sheet["A{}".format(active_row - 1)] = question_label

        # Add in MIP name (note assume Python 2 hence encode necessary)
        machines_sheet["B{}".format(active_row - 1)] = mip.encode()

        section_number += 1  # update sub-sub-sub-section for next question

    return active_row  # so know where to start at for end-of-tab note


def _main(args):
    """Main entry point.

    """
    # Defensive programming.
    if not os.path.exists(args.xls_template):
        raise ValueError("XLS template file does not exist")

    # Take generic template ready to process with institute-specific info.
    template_name = args.xls_template

    # Write out a customised template file for every institute.
    for institution in vocabs.get_institutes(args.institution_id):
        # !!! Must define this on each iteration instead of outside the 'for'
        # loop else iterations will include previous changes. !!!
        generic_template = load_workbook(filename=template_name)

        # Styles needed for clear, coordinated formatting of processed cells.
        generic_template.add_named_style(QUESTION_HEADER_STYLE)
        generic_template.add_named_style(QUESTION_INPUT_BOX_STYLE)
        
        # Customise the template appropriately to the given institute:
        #     1. Set the institute name
        set_institute_name_in_xls(institution, generic_template)

        #     2. Set the applicable CMIP6 models for this institute
        end_cell = set_applicable_models_in_xls(institution, generic_template)

        #     3. Set the applicable CMIP6 experiments for this institute
        tab_end_cell = set_applicable_experiments_in_xls(
            institution, generic_template, end_cell)

        # Add a note that the tab is done, with a reminder, for clarity.
        generic_template["Machine 1"]["B{}".format(tab_end_cell + 2)] = (
            "END of spreadsheet tab covering a single machine. Remember to "
            "document all CMIP6 machines for your institute in a separate tab."
        )
        generic_template["Machine 1"][
            "B{}".format(tab_end_cell + 2)
        ].font = Font(name="Helvetica Neue", size=14, italic=True, bold=True)
        generic_template["Machine 1"].row_dimensions[
            tab_end_cell + 2].height = 50
        generic_template["Machine 1"][
            "B{}".format(tab_end_cell + 2)
        ].alignment = Alignment(wrap_text=True)

        # Close template and write the customised template to a new XLS file.
        generic_template.close()
        final_spreadsheet_name = "{}_{}_machines.xlsx".format(
            constants.CMIP6_MIP_ERA, institution.canonical_name)
        generic_template.save(final_spreadsheet_name)

        # Place the template into the appropriate directory.
        # Write one file per institute.
        dest = io_mgr.get_machines_spreadsheet(institution)
        logger.log("moving xls file for {}".format(institution.raw_name))
        print(final_spreadsheet_name, dest)
        shutil.copy(final_spreadsheet_name, dest)


# Main entry point.
if __name__ == '__main__':
    _main(_ARGS.parse_args())
